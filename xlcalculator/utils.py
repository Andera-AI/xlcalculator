import collections
import re
from openpyxl.utils.cell import COORD_RE, SHEET_TITLE
from openpyxl.utils.cell import range_boundaries, get_column_letter
from enum import Enum

class TableReferenceError(Exception):
    """Base class for table reference errors"""
    pass

class InvalidTableReferenceError(TableReferenceError):
    """Raised when table reference syntax is invalid"""
    pass

class TableNotFoundError(TableReferenceError):
    """Raised when referenced table doesn't exist"""
    pass

class ColumnNotFoundError(TableReferenceError):
    """Raised when referenced column doesn't exist"""
    pass

class InvalidItemSpecifierError(TableReferenceError):
    """Raised when an invalid item specifier is used"""
    pass

MAX_COL = 18278
MAX_ROW = 1048576

# Item specifiers in Microsoft Excel structured references
class ItemSpecifier(str, Enum):
    All = "All"
    Headers = "Headers"
    Data = "Data"
    ThisRow = "This Row"
    Totals = "Totals" # not supported by our code yet

def resolve_sheet(sheet_str):
    sheet_str = sheet_str.strip()
    sheet_match = re.match(SHEET_TITLE.strip(), sheet_str + '!')
    if sheet_match is None:
        # Internally, sheets are not properly quoted, so consider the entire
        # string.
        return sheet_str

    return sheet_match.group("quoted") or sheet_match.group("notquoted")


def resolve_address(addr):
    # Addresses without sheet name are not supported.
    sheet_str, addr_str = addr.split('!')
    sheet = resolve_sheet(sheet_str)
    coord_match = COORD_RE.split(addr_str)
    col, row = coord_match[1:3]
    return sheet, col, row


def resolve_ranges(ranges, default_sheet='Sheet1', sheet_max_row=None):
    sheet = None
    range_cells = collections.defaultdict(set)
    for rng in ranges.split(','):
        # Handle sheets in range.
        if '!' in rng:
            sheet_str, rng = rng.split('!')
            rng_sheet = resolve_sheet(sheet_str)
            if sheet is not None and sheet != rng_sheet:
                raise ValueError(
                    f'Got multiple different sheets in ranges: '
                    f'{sheet}, {rng_sheet}'
                )
            sheet = rng_sheet
        min_col, min_row, max_col, max_row = range_boundaries(rng)

        # Unbound ranges (e.g., A:A) might not have these set! So use the max row of the sheet with data if available
        min_col = min_col or 1
        min_row = min_row or 1
        max_col = max_col or MAX_COL
        max_row = max_row or sheet_max_row or MAX_ROW

        # Excel ranges are boundaries inclusive!
        for row_idx in range(min_row or 1, max_row + 1):
            row_cells = range_cells[row_idx]
            for col_idx in range(min_col, max_col + 1):
                row_cells.add(col_idx)

    # Now convert the internal structure to a matrix of cell addresses.
    sheet = default_sheet if sheet is None else sheet
    sheet_str = sheet + '!' if sheet else ''
    return sheet, [
        [
            f'{sheet_str}{get_column_letter(col_idx)}{row_idx}'
            for col_idx in sorted(row_cells)
        ]
        for row_idx, row_cells in sorted(range_cells.items())
    ]

def resolve_table_ranges(ranges, tables: dict[str, any], cur_cell_addr: str | None = None):
    """
    Given a structured reference / table reference, return the cell range that it references in the format of "<sheet>!<range>"
    Documentation on syntax rules: https://support.microsoft.com/en-au/office/using-structured-references-with-excel-tables-f5ed2452-2337-4f71-bed3-c8ae6d2b276e
    
    Args:
        ranges: The table reference string (e.g., "Table1[Col1]")
        tables: Dictionary of XLTable objects
        cur_cell_addr: Current cell address for #This Row specifier
        
    Returns:
        str: Cell range in format "<sheet>!<range>"
    """
    try:
        # Parse the table reference
        table_range_components = _parse_table_range(ranges)
        table_specifier_components = _extract_table_specifiers(table_range_components["specifier"])
        
        # Process specifiers
        item_specifiers = []
        start_col = None
        end_col = None
        for table_specifier_component in table_specifier_components:
            start_col, end_col = _parse_specifier(table_specifier_component, item_specifiers)
            
        # Sanitize column names
        start_col = _sanitize_table_column_name(start_col) if start_col else start_col
        end_col = _sanitize_table_column_name(end_col) if end_col else end_col
        
        # Get and validate table range
        table_range = _get_table_range(
            table_range_components["table"], 
            start_col, 
            end_col, 
            item_specifiers, 
            tables, 
            cur_cell_addr
        )
        
        if not table_range:
            raise InvalidTableReferenceError("Unable to extract cell range of table range")
            
        return table_range
        
    except Exception as e:
        raise Exception(f"Error resolving table range: {e}")

def _parse_table_range(term: str) -> dict:
    """
    Given a potential structured reference / table reference, return the sheet, table, and specifier (whatever is inside the outermost [])
    Uses regex to parse the term

    Args:
        term: The table reference string to parse
        
    Returns:
        dict: Dictionary containing sheet, table, and specifier
    """
    match = re.match(
        r"""^(?:(?P<sheet>[^!\[\]]+)!){0,1}   # Optional 'Sheet!'
            (?P<table>[^\[\]]+)               # Table name
            \[(?P<specifier>.*)\]$            # Everything inside outermost []
        """,
        term,
        re.VERBOSE
    )
    if not match:
        raise InvalidTableReferenceError("Term doesn't follow structured reference pattern")
    
    specifier = match.group("specifier")
    if specifier is None:
        raise InvalidTableReferenceError("Unable to extract specifier from term")
    elif isinstance(specifier, str):
        specifier = specifier.strip()

    sheet = match.group("sheet")
    table = match.group("table")
    if table is None:
        raise InvalidTableReferenceError("Unable to extract table from term")

    return {
        "sheet": sheet,
        "table": table,
        "specifier": specifier,
    }

def _extract_table_specifiers(table_specifier: str) -> list[str]:
    """
    Given the main table specifier, return a list of specifiers.
    The main table specifier pattern is a list of [] separated by commas, 
    unless there is only 1 specifer, of which there might not be any []
    """
    parts = []
    depth = 0
    current = []

    for char in table_specifier:
        if char == '[':
            depth += 1
        elif char == ']':
            depth -= 1

        if char == ',' and depth == 0:
            part = ''.join(current).strip()
            if part.startswith('[') and part.endswith(']'):
                parts.append(part)
            current = []
        else:
            current.append(char)

    # Add last part
    part = ''.join(current).strip()
    if part.startswith('[') and part.endswith(']'):
        parts.append(part)

    if len(parts) == 0:
        return [table_specifier]
    return parts

def _parse_specifier(specifier: str, item_specifiers: list[str]) -> tuple[str, str]:
    """
    A specifier can be a 1) item specifier, 2) a column range, or 3) a single column
    """
    # Case 0: empty specifier
    if specifier == "":
        return None, None
    
    # Strip []
    if specifier.startswith("[") and specifier.endswith("]"):
        specifier = specifier[1:-1]

    # Case 1: item specifier
    if specifier[0] == "#":
        item_specifiers.append(specifier[1:]) 
        return None, None

    # Case 2: columnrange
    # We can assume from experiments that the range is always split by "]:[", with no space in between from excel
    if "]:[" in specifier:
        start_col, end_col = specifier.split("]:[")
        start_col = start_col
        end_col = end_col
        return start_col, end_col
    
    # Case 3: single column
    return specifier, specifier

def _sanitize_table_column_name(column_name: str) -> str:
    """
    Remove escape characters ' from the column name
    """
    special_chars = ['[', ']', "'", '#', '@']
    new_column_name = ""
    for i in range(0, len(column_name)):
        if column_name[i] == "'":
            if i < len(column_name) - 1 and column_name[i+1] in special_chars:
                continue
        new_column_name += column_name[i]
    return new_column_name

def _get_table_range(table_name: str, start_col: str | None, end_col: str | None, 
                    item_specifiers: list[str], tables: dict, cur_cell_addr: str | None = None) -> str | None:
    """
    Given a dictionary of tables, translate the start column, end column, and item specifiers to return the table cell range in the format of "<sheet>!<range>"
    
    Args:
        table_name: Name of the table to get range for
        start_col: Starting column name
        end_col: Ending column name
        item_specifiers: List of item specifiers (#Headers, #Data, etc.)
        tables: Dictionary of available tables
        cur_cell_addr: Current cell address for #This Row specifier
        
    Returns:
        str: Cell range in format "<sheet>!<range>"
    """
    table_name = _translate_table_name(table_name, tables)
    if end_col is None:
        end_col = start_col

    # Get table column range
    table_range = tables[table_name].cell_range
    min_col, min_row, max_col, max_row = range_boundaries(table_range)

    # Apply item specifiers, limiting the rows range
    if len(item_specifiers) == 0:
        # Empty specifier means the data part of the table
        min_row = min_row + tables[table_name].header_row_count
        max_row = max_row - 1 if tables[table_name].has_totals_row else max_row
    elif ItemSpecifier.All in item_specifiers:
        pass
    else:
        # Get the final bounds of the table range
        # There are some specifiers that are mutually exclusive, so we need to check for that (e.g. This Row, Headers & Totals cannot be used together)
        fixed_min_row = None
        fixed_max_row = None
        temp_min_row = max_row
        temp_max_row = min_row
        for item_specifier in item_specifiers:
            match item_specifier:
                case ItemSpecifier.Headers:
                    if fixed_max_row is not None:
                        raise InvalidTableReferenceError("Headers specifier cannot be used together with Headers / This Row specifiers")
                    fixed_min_row = min_row
                    temp_max_row = max(temp_max_row, min_row + tables[table_name].header_row_count - 1)
                case ItemSpecifier.Data:
                    # Data: No Headers and Totals Rows
                    temp_min_row = min(temp_min_row, min_row + tables[table_name].header_row_count)
                    temp_max_row = max(temp_max_row, max_row - 1 if tables[table_name].has_totals_row else max_row)
                case ItemSpecifier.ThisRow:
                    # Current row only
                    if cur_cell_addr is None:
                        raise InvalidTableReferenceError("Current cell address is not provided for #This Row item specifier")
                    if fixed_min_row is not None or fixed_max_row is not None:
                        raise InvalidTableReferenceError("This Row specifier cannot be used together with other specifiers")
                    coor = cur_cell_addr.rsplit("!", 1)[-1]
                    _, fixed_min_row, _, fixed_max_row = range_boundaries(coor)
                    break
                case ItemSpecifier.Totals:
                    if not tables[table_name].has_totals_row:
                        # Should return null ideally
                        raise InvalidTableReferenceError("Table does not have a totals row")
                    if fixed_min_row is not None:
                        raise InvalidTableReferenceError("Totals specifier cannot be used together with Headers / This Row specifiers")
                    fixed_max_row = max_row
                    temp_min_row = min(temp_min_row, max_row)
                case _:
                    raise InvalidItemSpecifierError(f"Item specifier not supported: {item_specifier}")
        min_row = temp_min_row if fixed_min_row is None else fixed_min_row
        max_row = temp_max_row if fixed_max_row is None else fixed_max_row

    # Special case: no column range specified, so span all columns
    if start_col is None:
        return f"{tables[table_name].sheet}!{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    # Get column range indexes from column names, limiting column range
    start_col_index = None
    end_col_index = None
    index = min_col

    column_name_count = {}
    for column in tables[table_name].columns:
        # Append counter to column name if it's not unique
        column_name = column.name
        if column_name in column_name_count:
            column_name_count[column_name] += 1
            column_name = f"{column_name}_{column_name_count[column_name]}"
        else:
            column_name_count[column_name] = 1
        if column_name == start_col:
            start_col_index = index
        if column_name == end_col:
            end_col_index = index
        index += 1

    if start_col_index is None:
        raise ColumnNotFoundError(f"Column '{start_col}' not found in table '{table_name}'")
    if end_col_index is None:
        raise ColumnNotFoundError(f"Column '{end_col}' not found in table '{table_name}'")
        
    return f"{tables[table_name].sheet}!{get_column_letter(start_col_index)}{min_row}:{get_column_letter(end_col_index)}{max_row}"

def _translate_table_name(table_name: str, tables: dict) -> str:
    """
    Since we're appending _<sheet_name> to the table name in the upload code, and while the excel formula still retains the original table name,
    we need to find the table name that matches the original table name.
    """
    if table_name in tables:
        return table_name
    else:
        for name, table in tables.items():
            if "_" in name:
                original_table_name, potential_sheet = name.split("_", 1)
                if original_table_name == table_name and table.sheet == potential_sheet:
                    return name
    raise TableNotFoundError(f"Table '{table_name}' not found. Available tables: {list(tables.keys())}")